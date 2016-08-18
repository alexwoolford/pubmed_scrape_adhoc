#!/usr/bin/env python

import openpyxl
from openpyxl.utils import get_column_letter


class ExcelDoctorExtract():

    def __init__(self):
        pass

    def parse(self, filename):
        wb = openpyxl.load_workbook(filename)
        sheet = wb.get_active_sheet()
        max_row = sheet.max_row
        max_column = sheet.max_column

        doctor_name_set = set()

        row = 1
        column = 1
        while row < max_row:
            while column < max_column:
                cell_value = sheet[get_column_letter(column) + str(row)].value
                try:
                    parsed_name = self.parse_doctor_name(cell_value)
                    if parsed_name:
                        doctor_name_set.add(parsed_name)
                except:
                    pass
                column += 1
            row += 1
            column = 1

        return doctor_name_set

    def parse_doctor_name(self, cell_value):
        if cell_value.count(",") >= 3:
            doctor_name = cell_value.split(",")[0]
            doctor_name = " ".join([doctor_name_fragment for doctor_name_fragment in doctor_name.split() if len(doctor_name_fragment) > 2])
            return doctor_name
        else:
            return None


if __name__ == "__main__":
    excel_doctor_extract = ExcelDoctorExtract()
    print excel_doctor_extract.parse('Faculty TOBI 2010-2016 & 2017.xlsx')
